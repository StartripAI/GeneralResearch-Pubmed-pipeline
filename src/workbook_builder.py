#!/usr/bin/env python3
"""Generic evidence workbook builder.

This script creates a styled multi-sheet Excel workbook from CSV inputs.
It is intentionally domain-agnostic and English-only for public distribution.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MCKINSEY_BLUE = "003A70"


def _display_width(value: object) -> int:
    text = str(value or "")
    return max(0, len(text))


def style_sheet(ws) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(fill_type="solid", fgColor=MCKINSEY_BLUE)
    header_font = Font(color="FFFFFF", bold=True)
    first_col_font = Font(color=MCKINSEY_BLUE, bold=True)
    base_font = Font(color="000000", bold=False)

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row <= 0 or max_col <= 0:
        return

    for col_idx in range(1, max_col + 1):
        max_w = 0
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            max_w = max(max_w, _display_width(cell.value))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_idx == 1:
                cell.font = first_col_font
            else:
                cell.font = base_font

        width = min(60, max(10, max_w + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"


def build_field_dictionary(frames: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    seen = set()
    for sheet, df in frames.items():
        for col in df.columns:
            key = (col,)
            if key in seen:
                continue
            seen.add(key)
            sample = ""
            series = df[col].astype(str).str.strip()
            series = series[series != ""]
            if not series.empty:
                sample = series.iloc[0][:160]
            rows.append(
                {
                    "field_name": col,
                    "description": "See project schema and evidence contracts.",
                    "appears_in": sheet,
                    "sample": sample,
                }
            )

    out = pd.DataFrame(rows)
    out = out.sort_values(["field_name", "appears_in"]).reset_index(drop=True)
    return out


def parse_sheet_spec(spec: str) -> Dict[str, str]:
    if "=" not in spec:
        raise ValueError(f"Invalid --sheet spec: {spec}. Expected NAME=/path/to/file.csv")
    name, path = spec.split("=", 1)
    name = name.strip()
    path = path.strip()
    if not name or not path:
        raise ValueError(f"Invalid --sheet spec: {spec}")
    return {"name": name, "path": path}


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a styled Excel workbook from CSV sheets.")
    parser.add_argument("--sheet", action="append", default=[], help="Sheet spec: NAME=/path/to/file.csv")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    if not args.sheet:
        raise SystemExit("At least one --sheet is required.")

    frames: Dict[str, pd.DataFrame] = {}
    for spec in args.sheet:
        pair = parse_sheet_spec(spec)
        path = Path(pair["path"]).expanduser().resolve()
        if not path.exists():
            raise SystemExit(f"Input file not found: {path}")
        frames[pair["name"]] = pd.read_csv(path)

    field_dict = build_field_dictionary(frames)

    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    sheet_order = ["field_dictionary", *frames.keys()]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        field_dict.to_excel(writer, sheet_name="field_dictionary", index=False)
        for name, df in frames.items():
            df.to_excel(writer, sheet_name=name, index=False)
        for name in sheet_order:
            style_sheet(writer.book[name])

    print(f"Wrote workbook: {output}")


if __name__ == "__main__":
    main()
